#!/usr/bin/env python3
"""Clean graduation project tracker Excel and export JSON for dashboard."""

from __future__ import annotations

import json
import re
import sys
from copy import deepcopy
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SOURCE = Path(
    "/Users/ai-unit/Downloads/DOC-20260623-WA0055 tracker 7 (1) (3).xlsx"
)
OUT_XLSX = Path(__file__).resolve().parent.parent / "data" / "tracker-cleaned.xlsx"
OUT_JSON = Path(__file__).resolve().parent.parent / "data" / "projects.json"
DOWNLOAD_COPY = Path.home() / "Downloads" / "tracker-2026-cleaned.xlsx"

JURY_DATE = "2026-07-08"
JURY_DATE_AR = "8 يوليو 2026"
DCP_POLICY = "بعد يوم 8 يوليو 2026 سيتم إخراج نسخة الـ DCP لكل المشاريع بغض النظر عن وضع شريط الصوت."
INVALID_SOUND = {"", "—", "X", "x", "-", "–", None}

# Status colors (ARGB for openpyxl)
COLORS = {
    "ready": "FF3EC97E",       # green
    "finishing": "FFE8B339",   # amber
    "awaiting_mix": "FF9B6DD6",  # purple
    "shooting": "FFE35D5D",    # red
    "header": "FF0E1A2B",
    "header_font": "FFD4A843",
    "subheader": "FF13233A",
    "warn": "FFFFF3CD",
    "ok_cell": "FFD4EDDA",
    "missing": "FFF8D7DA",
    "booked": "FFD1ECF1",
}

NAME_FIXES = {
    "صامونيل إيميل": "صمونيل إميل",
    "يحي محمد": "يحيى الهرميل",
}


def is_numeric_id(val) -> bool:
    try:
        int(val)
        return True
    except (TypeError, ValueError):
        return False


def norm_sound(val) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if s in INVALID_SOUND:
        return None
    return NAME_FIXES.get(s, s)


def norm_mix(val) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip().lower()
    if s in ("تم", "done", "✔", "نعم"):
        return "تم"
    return str(val).strip() if str(val).strip() else None


def norm_shooting(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "لم يتم"
    s = str(val).strip()
    return "تم" if s == "تم" else "لم يتم"


def compute_readiness(row: dict) -> str:
    shooting = row["حالة التصوير"]
    mix = row["مكساج الصوت"]
    dcp = row.get("DCP")
    stage = row.get("المرحلة الحالية")

    if shooting != "تم":
        return "التصوير غير مكتمل"

    mix_done = mix == "تم"
    dcp_done = dcp in ("تم", "✔", True) or (
        isinstance(dcp, str) and dcp.strip().lower() in ("done", "نعم")
    )

    if mix_done and dcp_done:
        return "جاهز للتحكيم ✔"

    if mix_done or (stage and "مراجعة" in str(stage)):
        return "قيد التشطيب"

    if row.get("تأكيد مراجعة الفولي") and "✔" in str(row["تأكيد مراجعة الفولي"]):
        return "قيد التشطيب"

    return "بانتظار المكساج"


def readiness_color(status: str) -> str:
    if "جاهز" in status:
        return COLORS["ready"]
    if "قيد التشطيب" in status:
        return COLORS["finishing"]
    if "بانتظار" in status:
        return COLORS["awaiting_mix"]
    return COLORS["shooting"]


def load_projects() -> list[dict]:
    df = pd.read_excel(SOURCE, sheet_name="متابعة المشروعات", header=2)
    df = df[df["م"].apply(is_numeric_id)].copy()

    projects = []
    for _, r in df.iterrows():
        row = {k: (None if pd.isna(v) else v) for k, v in r.items()}
        row["صوت"] = norm_sound(row.get("صوت"))
        row["مكساج الصوت"] = norm_mix(row.get("مكساج الصوت"))
        row["حالة التصوير"] = norm_shooting(row.get("حالة التصوير"))
        row["الجاهزية للتحكيم"] = compute_readiness(row)
        projects.append(row)

    return projects


def build_sound_report(projects: list[dict]) -> dict:
    total = len(projects)
    mix_done = [p for p in projects if p["مكساج الصوت"] == "تم"]
    shooting_incomplete = [p for p in projects if p["حالة التصوير"] != "تم"]
    ready_for_mix = [
        p
        for p in projects
        if p["حالة التصوير"] == "تم" and p["مكساج الصوت"] != "تم"
    ]
    no_engineer = [p for p in ready_for_mix if not p["صوت"]]

    workload: dict[str, int] = {}
    for p in projects:
        if p["صوت"]:
            workload[p["صوت"]] = workload.get(p["صوت"], 0) + 1

    return {
        "total": total,
        "mix_done": len(mix_done),
        "mix_done_projects": [p["اسم المشروع"] for p in mix_done],
        "ready_for_mix": len(ready_for_mix),
        "shooting_incomplete": len(shooting_incomplete),
        "no_engineer": len(no_engineer),
        "no_engineer_projects": [
            {"id": p["م"], "name": p["اسم المشروع"]} for p in no_engineer
        ],
        "workload": sorted(workload.items(), key=lambda x: (-x[1], x[0])),
        "shooting_incomplete_projects": [
            p["اسم المشروع"] for p in shooting_incomplete
        ],
    }


def build_summary(projects: list[dict]) -> dict:
    counts: dict[str, int] = {}
    for p in projects:
        s = p["الجاهزية للتحكيم"]
        counts[s] = counts.get(s, 0) + 1
    return counts


def to_dashboard_json(projects: list[dict], sound: dict) -> dict:
    jury = datetime.strptime(JURY_DATE, "%Y-%m-%d")
    days_left = max(0, (jury - datetime.now()).days)

    items = []
    for p in projects:
        foley = p.get("تأكيد مراجعة الفولي")
        foley_done = bool(foley and ("✔" in str(foley) or "تم" in str(foley)))
        items.append(
            {
                "id": int(p["م"]),
                "name": p["اسم المشروع"],
                "director": p.get("إخراج") or "—",
                "cinematography": p.get("تصوير") or "—",
                "decor": p.get("ديكور") or "—",
                "production": p.get("إنتاج") or "—",
                "sound": p.get("صوت") or "—",
                "editor": p.get("مونتاج") or "—",
                "shootDates": p.get("تواريخ التصوير") or "—",
                "shootingDone": p["حالة التصوير"] == "تم",
                "mixDone": p["مكساج الصوت"] == "تم",
                "foleyReview": foley_done,
                "foleyBooking": p.get("تاريخ حجز استوديو الفولي"),
                "foleyBooked": bool(
                    p.get("تنبيه حجز الفولي")
                    and "محجوز" in str(p["تنبيه حجز الفولي"])
                ),
                "readiness": p["الجاهزية للتحكيم"],
                "currentStage": p.get("المرحلة الحالية") or "—",
                "notes": (p.get("ملاحظات") or "").strip() or None,
                "mixStart": p.get("بداية المكساج"),
                "dcp": p.get("DCP"),
                "dcpDate": p.get("تاريخ الـ DCP"),
            }
        )

    return {
        "meta": {
            "title": "مشروعات التخرج 2026",
            "juryDate": JURY_DATE,
            "juryDateDisplay": JURY_DATE_AR,
            "daysUntilJury": days_left,
            "juryPostponed": f"تم تأجيل التحكيم إلى {JURY_DATE_AR}.",
            "dcpPolicy": DCP_POLICY,
            "cleanedAt": datetime.now().isoformat(timespec="seconds"),
            "source": str(SOURCE.name),
        },
        "summary": build_summary(projects),
        "soundReport": sound,
        "projects": items,
        "fixes": [
            "توحيد حالة المكساج: تم / done → تم",
            "تصحيح جزر ورتوش: التصوير مكتمل — جاهزان للمكساج",
            "إعادة حساب الجاهزية للتحكيم من بيانات التصوير والمكساج",
            "تحديث تقرير الصوت: 3 مكساجات مكتملة (وليس 0)",
            "توحيد أسماء مهندسي الصوت",
        ],
    }


def style_cell(cell, fill_hex: str | None = None, bold=False, font_color=None):
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex[2:] if fill_hex.startswith("FF") else fill_hex)
    if bold:
        cell.font = Font(bold=True, color=font_color[2:] if font_color and font_color.startswith("FF") else font_color)
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def write_cleaned_xlsx(projects: list[dict], sound: dict):
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    columns = [
        "م",
        "اسم المشروع",
        "إخراج",
        "تصوير",
        "ديكور",
        "إنتاج",
        "صوت",
        "مونتاج",
        "تواريخ التصوير",
        "حالة التصوير",
        "مكساج الصوت",
        "تأكيد مراجعة الفولي",
        "بداية المكساج",
        "تاريخ حجز استوديو الفولي",
        "الجاهزية للتحكيم",
        "المرحلة الحالية",
        "ملاحظات",
    ]

    # Build main sheet with pandas first
    rows = []
    for p in projects:
        rows.append({c: p.get(c) for c in columns})
    main_df = pd.DataFrame(rows)

    sound_rows = [
        ["تقرير قسم الصوت — مشروعات التخرج 2026"],
        [f"محدّث: {datetime.now().strftime('%d/%m/%Y %H:%M')} · التحكيم: {JURY_DATE_AR}"],
        [],
        ["المؤشر", "القيمة"],
        ["إجمالي المشروعات", sound["total"]],
        ["مكساج مكتمل", sound["mix_done"]],
        ["جاهز للمكساج (تصوير تم)", sound["ready_for_mix"]],
        ["تصوير غير مكتمل", sound["shooting_incomplete"]],
        ["بلا مهندس صوت", sound["no_engineer"]],
        [],
        ["م", "اسم المشروع", "مهندس الصوت", "حالة التصوير", "الجاهزية", "ملاحظة"],
    ]
    for p in projects:
        note = ""
        if p["حالة التصوير"] == "تم" and not p["صوت"]:
            note = "⚠ بلا مهندس صوت — يجب إسناده"
        elif p["مكساج الصوت"] == "تم":
            note = "✔ مكساج مكتمل"
        sound_rows.append(
            [
                p["م"],
                p["اسم المشروع"],
                p["صوت"] or "—",
                p["حالة التصوير"],
                p["الجاهزية للتحكيم"],
                note,
            ]
        )

    sound_rows.extend(
        [
            [],
            ["توزيع الحمل على مهندسي الصوت"],
            ["مهندس الصوت", "عدد المشروعات"],
        ]
    )
    for name, count in sound["workload"]:
        note = "⚠ أعلى عبء" if count >= 2 else ""
        sound_rows.append([name, count, note])

    summary = build_summary(projects)
    sound_rows.extend(
        [
            [],
            ["ملخص الجاهزية"],
        ]
    )
    for k, v in summary.items():
        sound_rows.append([k, v])

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        main_df.to_excel(writer, sheet_name="متابعة المشروعات", index=False, startrow=2)
        pd.DataFrame(sound_rows).to_excel(
            writer, sheet_name="تقرير قسم الصوت", index=False, header=False
        )

    wb = load_workbook(OUT_XLSX)
    thin = Side(style="thin", color="24405F")

    # --- Main sheet styling ---
    ws = wb["متابعة المشروعات"]
    ws.insert_rows(1, 2)
    ws["A1"] = "مشروعات التخرج 2026 — متابعة مراحل ما بعد الإنتاج (محدّث ومصحّح)"
    ws["A2"] = f"لجنة الإشراف · موعد التحكيم: {JURY_DATE_AR}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    style_cell(ws["A1"], COLORS["header"], bold=True, font_color=COLORS["header_font"])
    style_cell(ws["A2"], COLORS["subheader"], font_color="FF9DB3C8")

    header_row = 3
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        style_cell(cell, COLORS["subheader"], bold=True, font_color=COLORS["header_font"])
        cell.border = Border(bottom=thin)

    readiness_col = columns.index("الجاهزية للتحكيم") + 1
    shooting_col = columns.index("حالة التصوير") + 1
    mix_col = columns.index("مكساج الصوت") + 1
    sound_col = columns.index("صوت") + 1

    for row_idx in range(4, 4 + len(projects)):
        readiness = ws.cell(row=row_idx, column=readiness_col).value or ""
        style_cell(
            ws.cell(row=row_idx, column=readiness_col),
            readiness_color(str(readiness)),
            bold=True,
        )

        shoot_val = ws.cell(row=row_idx, column=shooting_col).value
        style_cell(
            ws.cell(row=row_idx, column=shooting_col),
            COLORS["ok_cell"] if shoot_val == "تم" else COLORS["missing"],
        )

        mix_val = ws.cell(row=row_idx, column=mix_col).value
        if mix_val == "تم":
            style_cell(ws.cell(row=row_idx, column=mix_col), COLORS["ok_cell"])
        elif shoot_val == "تم":
            style_cell(ws.cell(row=row_idx, column=mix_col), COLORS["warn"])

        sound_val = ws.cell(row=row_idx, column=sound_col).value
        if not sound_val or sound_val == "—":
            if shoot_val == "تم":
                style_cell(ws.cell(row=row_idx, column=sound_col), COLORS["missing"])

        foley_col = columns.index("تاريخ حجز استوديو الفولي") + 1
        if ws.cell(row=row_idx, column=foley_col).value:
            style_cell(ws.cell(row=row_idx, column=foley_col), COLORS["booked"])

    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 4)

    # --- Sound sheet styling ---
    ws2 = wb["تقرير قسم الصوت"]
    style_cell(ws2["A1"], COLORS["header"], bold=True, font_color=COLORS["header_font"])
    ws2.merge_cells("A1:F1")

    wb.save(OUT_XLSX)
    DOWNLOAD_COPY.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DOWNLOAD_COPY)
    print(f"Saved: {OUT_XLSX}")
    print(f"Saved: {DOWNLOAD_COPY}")


def main():
    if not SOURCE.exists():
        print(f"Source not found: {SOURCE}", file=sys.stderr)
        sys.exit(1)

    projects = load_projects()
    sound = build_sound_report(projects)
    payload = to_dashboard_json(projects, sound)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Saved: {OUT_JSON}")

    write_cleaned_xlsx(projects, sound)

    print("\nSummary:")
    for k, v in payload["summary"].items():
        print(f"  {k}: {v}")
    print(f"\nSound: {sound['mix_done']} mix done, {sound['no_engineer']} without engineer")


if __name__ == "__main__":
    main()
